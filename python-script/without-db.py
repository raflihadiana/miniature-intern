from pyxlsb import open_workbook as open_xlsb
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
import time

# Set the path to the directory containing the .xlsb files
xlsb_path = input("Enter the path to the directory containing the .xlsb files: ")

# Loop over the .xlsb files in the directory
for filename in os.listdir(xlsb_path):
    if filename.endswith('.xlsb'):
        # Get user input for week and year
        week = input(f"Enter week value for {filename}: ")
        year = int(input(f"Enter year value for {filename}: "))

        # Convert .xlsb file to .xlsx format
        xlsb_file_path = os.path.join(xlsb_path, filename)
        xlsx_file_path = os.path.join(xlsb_path, f"{os.path.splitext(filename)[0]}.xlsx")
        print(f"Converting {xlsb_file_path} to {xlsx_file_path}...")
        start_time = time.time()
        try:
            sheet_to_convert = "4G NQI Capacity Cell"
            workbook = open_xlsb(xlsb_file_path)
            wb = Workbook()
            worksheet = None
            for sheetname in workbook.sheets:
                if sheetname == sheet_to_convert:
                    worksheet = wb.create_sheet(title=sheetname)
                    for row in workbook.get_sheet(sheetname):
                        worksheet.append([cell.v for cell in row])

            if worksheet:
                wb.save(xlsx_file_path)
            else:
                print(f"Sheet {sheet_to_convert} not found in {xlsb_file_path}")

        except Exception as e:
            print(f"Error converting {xlsb_file_path} to {xlsx_file_path}: {str(e)}")

        end_time = time.time()
        duration = end_time - start_time
        print(f"{xlsb_file_path} converted to {xlsx_file_path} in {duration:.2f} seconds.")

        # Load the .xlsx file into a pandas dataframe
        print(f"Loading file {xlsx_file_path}...")
        start_time = time.time()
        try:
            # Handle locked files by retrying for up to 5 minutes
            for i in range(60):
                try:
                    wb = load_workbook(xlsx_file_path, read_only=True)
                    break
                except PermissionError:
                    if i == 0:
                        print(f"Warning: file {xlsx_file_path} is locked, retrying in 5 seconds...")
                    time.sleep(5)
            else:
                raise Exception(f"Could not open file {xlsx_file_path} because it is locked.")
                
            sheet = wb[sheet_to_convert]
            data = []
            for row in sheet.rows:
                data.append([cell.value for cell in row])

            if len(data) > 1:
                excel_file = pd.DataFrame(data[1:], columns=data[0])
                excel_file.columns = excel_file.iloc[0]
                excel_file.columns.name=None
                excel_file.drop(0, axis=0, inplace=True)
                excel_file.insert(0, 'Year', year)
                excel_file.insert(1, 'Week', week)
            else:
                print(f"Warning: sheet {sheet.title} in file {xlsx_file_path} is empty.")
                excel_file = pd.DataFrame(columns=['year', 'week'])

            # Modify the data in the dataframe as needed
            print(f"Modifying data in file {xlsx_file_path}...")
            excel_file = excel_file.replace(to_replace=r'^\d*\.\d*ENULL$', value=' ', regex=True)
            excel_file = excel_file.replace(['#N/A'], ' ')
            excel_file = excel_file.replace(['#DIV/0!'], ' ')
            excel_file = excel_file.replace(['NA'], ' ')

            # Convert the dataframe to CSV and save to a file with a new name
            new_filename = os.path.splitext(filename)[0] + f'_{sheet_to_convert}_modified.csv'
            csv_file_path = os.path.join(xlsb_path, new_filename)
            print(f"Saving file {csv_file_path}...")
            excel_file.to_csv(csv_file_path, index=False)
            
        except Exception as e:
            print(f"Error processing {xlsx_file_path}: {str(e)}")

        finally:
            if 'wb' in locals():
                wb.close()

        end_time = time.time()
        duration = end_time - start_time
        print(f"File {filename} processed in {duration:.2f} seconds.")

print("All files processed.")